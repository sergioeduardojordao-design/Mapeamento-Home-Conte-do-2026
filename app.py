import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font
import os
import time
import folium
from streamlit_folium import st_folium
import random

# Cores oficiais da Home Conteúdo
HEX_HOME = "#E47673"

st.set_page_config(
    page_title="Home Conteúdo - Dashboard",
    page_icon="🏠",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Estilização do App Tech
st.markdown(f"""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght=400;600;700&display=swap');
    html, body, [class*="css"] {{ font-family: 'Inter', sans-serif; background-color: #f8f9fa; }}
    .stTabs [data-baseweb="tab-panel"] {{
        background: rgba(255, 255, 255, 0.8); border-radius: 20px; padding: 25px;
        border: 1px solid rgba(255, 255, 255, 0.3); box-shadow: 0 8px 32px 0 rgba(31, 38, 135, 0.05);
        backdrop-filter: blur(10px);
    }}
    .stButton>button {{
        background-color: {HEX_HOME}; color: white; border-radius: 12px; border: none;
        padding: 12px 24px; font-weight: 600; box-shadow: 0 4px 15px rgba(228, 118, 115, 0.3);
        transition: all 0.3s ease; width: 100%;
    }}
    .stButton>button:hover {{ transform: translateY(-2px); background-color: #d16260; }}
    .stDataFrame {{ border-radius: 15px; overflow: hidden; border: 1px solid #eee; }}
    .logo-container {{ display: flex; justify-content: center; align-items: center; flex-direction: column; padding: 20px; }}
    </style>
""", unsafe_allow_html=True)

EXCEL_FILE = "banco_criadoras.xlsx"

COORDENADAS_UF = {
    "SP": (-23.5505, -46.6333), "RJ": (-22.9068, -43.1729), "MG": (-19.9167, -43.9345), "ES": (-20.2976, -40.2958),
    "PR": (-25.4284, -49.2733), "RS": (-30.0346, -51.2177), "SC": (-27.5954, -48.5480), "DF": (-15.7801, -47.9292),
    "GO": (-16.6869, -49.2648), "MT": (-15.6010, -56.0974), "MS": (-20.4428, -54.6464), "BA": (-12.9714, -38.5014),
    "PE": (-8.0543, -34.8813), "CE": (-3.7173, -38.5434), "MA": (-2.5297, -44.3028), "AL": (-9.6658, -35.7350),
    "PB": (-7.1150, -34.8631), "PI": (-5.0920, -42.8038), "RN": (-5.7945, -35.2110), "SE": (-10.9111, -37.0717),
    "AM": (-3.1190, -60.0217), "PA": (-1.4558, -48.4902), "RO": (-8.7612, -63.9039), "AP": (0.0340, -51.0694),
    "TO": (-10.1674, -48.3277), "AC": (-9.9740, -67.8076), "RR": (2.8197, -60.6732)
}

def obter_coordenadas(uf):
    lat, lng = COORDENADAS_UF.get(str(uf).strip().upper(), (-14.2350, -51.9253))
    ajuste_lat = random.uniform(-0.15, 0.15)
    ajuste_lng = random.uniform(-0.15, 0.15)
    return lat + ajuste_lat, lng + ajuste_lng

def limpar_valor_numerico(val):
    if pd.isna(val) or str(val).strip() == "" or str(val).startswith('='): return 0
    try:
        if isinstance(val, (int, float)): return int(val)
        return int(float(str(val).strip().replace(" ", "").replace(".", "").replace(",", "")))
    except: return 0

def limpar_taxa_float(val):
    if pd.isna(val) or str(val).strip() == "" or str(val).startswith('='): return 0.0
    try:
        if isinstance(val, (int, float)): return float(val)
        txt = str(val).replace("%", "").strip().replace(",", ".")
        return float(txt)
    except: return 0.0

def descobrir_linha_cabecalho():
    if not os.path.exists(EXCEL_FILE): return 1
    df_bruto = pd.read_excel(EXCEL_FILE, header=None)
    for i, row in df_bruto.iterrows():
        if "CRIADORA" in [str(x).upper().strip() for x in row.values]:
            return i
    return 0

def achar_ultima_linha_com_texto(ws):
    for row in range(ws.max_row, 0, -1):
        celula = ws.cell(row=row, column=3).value
        if celula is not None and str(celula).strip() != "" and str(celula).upper() != "CRIADORA":
            return row
    return ws.max_row

def carregar_e_processar_dados():
    if not os.path.exists(EXCEL_FILE): return pd.DataFrame()
    try:
        linha_idx = descobrir_linha_cabecalho()
        
        wb_links = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
        ws_links = wb_links.active
        
        wb_dados = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        ws_dados = wb_dados.active
        
        colunas_nomes = [str(ws_dados.cell(linha_idx + 1, col).value).strip() for col in range(1, ws_dados.max_column + 1)]
        
        dados = []
        for r_idx in range(linha_idx + 2, ws_dados.max_row + 1):
            if str(ws_dados.cell(r_idx, 3).value).strip() in ["", "None", "CRIADORA"]: continue
            linha = {}
            for c_idx, c_name in enumerate(colunas_nomes, start=1):
                if c_name == "None" or c_name == "": continue
                
                cell_dados = ws_dados.cell(r_idx, c_idx)
                cell_links = ws_links.cell(r_idx, c_idx)
                
                if c_name in ["YOUTUBE", "INSTAGRAM", "TIK TOK", "FACEBOOK", "KWAI"] and cell_links.hyperlink:
                    linha[c_name] = cell_links.hyperlink.target
                else:
                    linha[c_name] = cell_dados.value if cell_dados.value is not None else ""
            dados.append(linha)
            
        df = pd.DataFrame(dados)
        
        if "PORTE" in df.columns:
            df["PERFIL"] = df["PORTE"].astype(str).str.strip().replace(["", "None", "nan", "0", "0.0"], "Não Definido")
        else:
            df["PERFIL"] = "Não Definido"
            
        colunas_texto_obrigatorias = ["CRIADORA", "CIDADE", "UF", "PRINCIPAL REDE", "E-mail de contato", "Telefone", "OBS", "COMPLEMENTO", "PERFIL"]
        for col_txt in colunas_texto_obrigatorias:
            if col_txt in df.columns:
                df[col_txt] = df[col_txt].fillna("").astype(str).str.strip().replace(["None", "nan", "0", "0.0"], "")

        df["Link YouTube"] = df["YOUTUBE"].fillna("").astype(str) if "YOUTUBE" in df.columns else ""
        df["Link Instagram"] = df["INSTAGRAM"].fillna("").astype(str) if "INSTAGRAM" in df.columns else ""
        df["Link TikTok"] = df["TIK TOK"].fillna("").astype(str) if "TIK TOK" in df.columns else ""
        
        for col in df.columns:
            if col in ["YOUTUBE", "INSTAGRAM", "TIK TOK", "CRIADORA", "CIDADE", "UF", "COMPLEMENTO", "E-mail de contato", "Telefone", "OBS", "PORTE", "PERFIL", "PRINCIPAL REDE", "Link YouTube", "Link Instagram", "Link TikTok"]:
                continue
            if "TAXA" not in col.upper() and "ENGAJAMENTO" not in col.upper():
                df[col] = df[col].apply(limpar_valor_numerico)

        # RECALCULO DAS TAXAS
        for idx, row in df.iterrows():
            ig_seg = limpar_valor_numerico(row.get("SEGUIDORES INSTAGRAM", 0))
            ig_interacoes = limpar_valor_numerico(row.get("VIEWS INSTAGRAM", 0)) + limpar_valor_numerico(row.get("CURTIDAS INSTAGRAM", 0)) + limpar_valor_numerico(row.get("COMENT. INSTAGRAM", 0))
            taxa_ig_val = float(ig_interacoes / ig_seg) if ig_seg > 0 else 0.0
            for c_taxa in df.columns:
                if "INSTAGRAM" in c_taxa.upper() and ("TAXA" in c_taxa.upper() or "ENGAJAMENTO" in c_taxa.upper()):
                    df.at[idx, c_taxa] = taxa_ig_val

            yt_seg = limpar_valor_numerico(row.get("INSCRITOS YOUTUBE", 0))
            yt_interacoes = limpar_valor_numerico(row.get("VIEWS YOUTUBE", 0)) + limpar_valor_numerico(row.get("CURTIDAS YOUTUBE", 0)) + limpar_valor_numerico(row.get("COMENT YOUTUBE", 0))
            taxa_yt_val = float(yt_interacoes / yt_seg) if yt_seg > 0 else 0.0
            for c_taxa in df.columns:
                if "YOUTUBE" in c_taxa.upper() and ("TAXA" in c_taxa.upper() or "ENGAJAMENTO" in c_taxa.upper()):
                    df.at[idx, c_taxa] = taxa_yt_val
            
            tt_seg = limpar_valor_numerico(row.get("SEGUIDORES TIK TOK", 0))
            tt_interacoes = limpar_valor_numerico(row.get("VIEWS TIK TOK", 0)) + limpar_valor_numerico(row.get("CURTIDAS TIK TOK", 0)) + limpar_valor_numerico(row.get("COMENTÁRIOS TIK TOK", 0))
            taxa_tt_val = float(tt_interacoes / tt_seg) if tt_seg > 0 else 0.0
            for c_taxa in df.columns:
                if "TIK" in c_taxa.upper() and ("TAXA" in c_taxa.upper() or "ENGAJAMENTO" in c_taxa.upper()):
                    df.at[idx, c_taxa] = taxa_tt_val
        
        for c_taxa in df.columns:
            if "TAXA" in c_taxa.upper() or "ENGAJAMENTO" in c_taxa.upper():
                df[c_taxa] = df[c_taxa].apply(limpar_taxa_float).astype(float)
                    
        return df
    except Exception as e:
        st.error(f"Erro ao processar dados da planilha: {e}")
        return pd.DataFrame()

def salvar_novo_registro_excel(dados):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        linha_idx = descobrir_linha_cabecalho()
        colunas_nomes = [str(ws.cell(linha_idx + 1, col).value).strip() for col in range(1, ws.max_column + 1)]
        
        proxima_linha = achar_ultima_linha_com_texto(ws) + 1
        font_link = Font(color="0563C1", underline="single")
        
        for idx, col_name in enumerate(colunas_nomes, start=1):
            if col_name == "CRIADORA": ws.cell(proxima_linha, idx, dados['nome'])
            elif col_name == "OBS": ws.cell(proxima_linha, idx, dados['obs'])
            elif col_name == "CIDADE": ws.cell(proxima_linha, idx, dados['cidade'])
            elif col_name == "UF": ws.cell(proxima_linha, idx, dados['uf'])
            elif col_name == "PORTE": ws.cell(proxima_linha, idx, dados['porte_manual'])
            elif col_name == "PRINCIPAL REDE": ws.cell(proxima_linha, idx, dados['rede_manual'])
            
            elif col_name == "YOUTUBE" and dados['rede_detectada'] == "YouTube":
                c = ws.cell(proxima_linha, idx, dados['nome'])
                c.hyperlink = dados['link_principal']; c.font = font_link
            elif col_name == "INSCRITOS YOUTUBE" and dados['rede_detectada'] == "YouTube":
                ws.cell(proxima_linha, idx, dados['seguidores'])
                
            elif col_name == "INSTAGRAM" and dados['rede_detectada'] == "Instagram":
                c = ws.cell(proxima_linha, idx, dados['nome'])
                c.hyperlink = dados['link_principal']; c.font = font_link
            elif col_name == "SEGUIDORES INSTAGRAM" and dados['rede_detectada'] == "Instagram":
                ws.cell(proxima_linha, idx, dados['seguidores'])
            elif col_name == "VIEWS INSTAGRAM" and dados['rede_detectada'] == "Instagram" and dados['teve_engajamento']:
                ws.cell(proxima_linha, idx, dados['v'])
            elif col_name == "CURTIDAS INSTAGRAM" and dados['rede_detectada'] == "Instagram" and dados['teve_engajamento']:
                ws.cell(proxima_linha, idx, dados['c'])
            elif col_name == "COMENT. INSTAGRAM" and dados['rede_detectada'] == "Instagram" and dados['teve_engajamento']:
                ws.cell(proxima_linha, idx, dados['m'])
            elif col_name == "ENGAJAMENTO TAXA INSTAGRAM" and dados['rede_detectada'] == "Instagram" and dados['teve_engajamento']:
                ws.cell(proxima_linha, idx, dados['engajamento'] / 100)
                
            elif col_name == "TIK TOK" and dados['rede_detectada'] == "TikTok":
                c = ws.cell(proxima_linha, idx, dados['nome'])
                c.hyperlink = dados['link_principal']; c.font = font_link
            elif col_name == "SEGUIDORES TIK TOK" and dados['rede_detectada'] == "TikTok":
                ws.cell(proxima_linha, idx, dados['seguidores'])
                
        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        st.error(f"Erro ao salvar: {e}")
        return False

# --- CONTROLE DE TELAS ---
if 'page' not in st.session_state: st.session_state.page = 'entrada'

if st.session_state.page == 'entrada':
    st.markdown("<div class='logo-container'>", unsafe_allow_html=True)
    if os.path.exists("logotipo.png"): st.image("logotipo.png", width=280)
    st.title("🏠 Home Conteúdo - Mapeamento Estratégico")
    st.subheader("Plataforma de inteligência e planejamento de mídia com influenciadoras")
    st.write("Escolha seu ambiente de trabalho:")
    c_l, c_r = st.columns(2)
    if c_l.button("👥 ACESSO EQUIPE (ADMIN)"): st.session_state.page = 'admin'; st.rerun()
    if c_r.button("🔍 ACESSO VISITANTE"): st.session_state.page = 'visitante'; st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)

elif st.session_state.page == 'admin':
    df_base = carregar_e_processar_dados()
    st.title("Painel Administrativo - Home Conteúdo")
    if st.button("⬅ Sair / Voltar para Início"): st.session_state.page = 'entrada'; st.rerun()
    
    if df_base.empty:
        st.error("Planilha 'banco_criadoras.xlsx' vazia ou não localizada na pasta.")
    else:
        tab1, tab2, tab3 = st.tabs(["📊 Inteligência de Dados", "➕ Novo Cadastro", "📈 Relatórios"])
        
        with tab1:
            with st.expander("🛠️ Filtros de Pesquisa Avançada", expanded=True):
                f1, f2, f3, f4 = st.columns(4)
                busca_txt = f1.text_input("Buscar por qualquer termo (Nome, Cidade, UF...):", "")
                
                porte_ops = sorted(list(df_base["PERFIL"].dropna().astype(str).unique())) if "PERFIL" in df_base.columns else []
                porte_f = f2.multiselect("Filtrar por Perfil:", porte_ops)
                
                rede_ops = sorted(list(df_base["PRINCIPAL REDE"].dropna().astype(str).unique())) if "PRINCIPAL REDE" in df_base.columns else []
                rede_f = f3.multiselect("Filtrar por Rede Principal:", rede_ops)
                
                uf_ops = sorted(list(df_base["UF"].dropna().astype(str).unique())) if "UF" in df_base.columns else []
                uf_f = f4.multiselect("Filtrar por Estado (UF):", uf_ops)
            
            df_f = df_base.copy()
            
            if busca_txt:
                colunas_para_buscar = [c for c in df_f.columns if c in ["CRIADORA", "CIDADE", "UF", "PERFIL", "PRINCIPAL REDE", "OBS", "E-mail de contato", "Telefone"]]
                mascara_global = df_f[colunas_para_buscar].astype(str).apply(lambda row: row.str.contains(busca_txt, case=False).any(), axis=1)
                df_f = df_f[mascara_global]
                
            if porte_f: df_f = df_f[df_f["PERFIL"].isin(porte_f)]
            if rede_f: df_f = df_f[df_f["PRINCIPAL REDE"].isin(rede_f)]
            if uf_f: df_f = df_f[df_f["UF"].isin(uf_f)]
            
            sub_visualizacao = st.radio("Formato de Exibição:", ["📋 Lista Dinâmica", "🗺️ Mapa Georreferenciado"], horizontal=True)
            
            if sub_visualizacao == "📋 Lista Dinâmica":
                st.markdown(f"**Total Encontrado:** {len(df_f)} influenciadoras")
                
                colunas_finais = ["CRIADORA", "PERFIL", "PRINCIPAL REDE", "CIDADE", "UF", "Link Instagram", "Link YouTube", "Link TikTok"]
                colunas_ignorar = ["YOUTUBE", "INSTAGRAM", "TIK TOK", "PORTE"]
                colunas_restantes = [c for c in df_f.columns if c not in colunas_finais and c not in colunas_ignorar]
                ordem_completa_colunas = colunas_finais + colunas_restantes
                
                df_render = df_f[ordem_completa_colunas].copy()
                for c_str in colunas_finais:
                    if c_str in df_render.columns:
                        df_render[c_str] = df_render[c_str].fillna("").astype(str)

                config_colunas = {
                    "CRIADORA": st.column_config.TextColumn("CRIADORA"),
                    "CIDADE": st.column_config.TextColumn("CIDADE"),
                    "UF": st.column_config.TextColumn("UF"),
                    "PERFIL": st.column_config.TextColumn("PERFIL"),
                    "PRINCIPAL REDE": st.column_config.TextColumn("PRINCIPAL REDE"),
                    "Link Instagram": st.column_config.LinkColumn("📸 INSTAGRAM", display_text="Abrir IG"),
                    "Link YouTube": st.column_config.LinkColumn("📺 YOUTUBE", display_text="Abrir YT"),
                    "Link TikTok": st.column_config.LinkColumn("🎵 TIKTOK", display_text="Abrir TT")
                }
                
                for col in ordem_completa_colunas:
                    if col in config_colunas: continue
                    if "TAXA" in col.upper() or "ENGAJAMENTO" in col.upper():
                        config_colunas[col] = st.column_config.NumberColumn(col, format="%.2f%%")
                    elif col in ["E-mail de contato", "Telefone", "OBS", "COMPLEMENTO", "FACEBOOK", "KWAI"]:
                        config_colunas[col] = st.column_config.TextColumn(col)
                        df_render[col] = df_render[col].fillna("").astype(str)
                    else:
                        config_colunas[col] = st.column_config.NumberColumn(col, format="%d")
                
                st.dataframe(df_render, use_container_width=True, hide_index=True, column_config=config_colunas)
            
            elif sub_visualizacao == "🗺️ Mapa Georreferenciado":
                st.markdown(f"**Pins Ativos:** Geolocalizando {len(df_f)} criadoras regionais.")
                m_map = folium.Map(location=[-14.2350, -51.9253], zoom_start=4)
                
                for _, row in df_f.iterrows():
                    uf_val = str(row.get("UF", "SP"))
                    lat, lng = obter_coordenadas(uf_val)
                    nome = str(row.get("CRIADORA", "Criadora"))
                    perfil = str(row.get("PERFIL", "Não Definido"))
                    cidade = str(row.get("CIDADE", ""))
                    email = str(row.get("E-mail de contato", "Não informado"))
                    tel = str(row.get("Telefone", "Não informado"))
                    obs = str(row.get("OBS", "-"))
                    
                    lk_ig = row.get("Link Instagram", "")
                    lk_yt = row.get("Link YouTube", "")
                    lk_tt = row.get("Link TikTok", "")
                    
                    # HTML da caixa (Popup maior)
                    popup_html = f"""
                    <div style="width: 250px; font-family: 'Inter', sans-serif; font-size: 12px; line-height: 1.5;">
                        <h4 style="margin: 0; color: {HEX_HOME};">{nome}</h4>
                        <hr style="margin: 5px 0;">
                        <b>Perfil:</b> {perfil}<br>
                        <b>Local:</b> {cidade}/{uf_val}<br>
                        <b>Contato:</b> {email}<br>
                        <b>Tel:</b> {tel}<br>
                        <div style="margin-top: 8px;">
                            {'<a href="'+lk_ig+'" target="_blank">📸 IG</a> ' if lk_ig else ''}
                            {'<a href="'+lk_yt+'" target="_blank">📺 YT</a> ' if lk_yt else ''}
                            {'<a href="'+lk_tt+'" target="_blank">🎵 TT</a>' if lk_tt else ''}
                        </div>
                        <div style="margin-top: 8px; font-style: italic; color: #555;">
                            <b>OBS:</b> {obs}
                        </div>
                    </div>
                    """
                    
                    folium.Marker(
                        location=[lat, lng], 
                        popup=folium.Popup(popup_html, max_width=300), 
                        icon=folium.Icon(color="red", icon="user", prefix="fa")
                    ).add_to(m_map)
                        
                st_folium(m_map, width=1200, height=520, returned_objects=[])

        with tab2:
            st.subheader("📝 Cadastrar Nova Influenciadora no Excel")
            c1, c2, c3 = st.columns([2, 3, 2])
            nome_cria = c1.text_input("Nome da Criadora *", key="cad_nome")
            link_cria = c2.text_input("Link da Rede Social Principal *", placeholder="https://instagram.com/perfil", key="cad_link")
            seg_cria = c3.number_input("Total de Seguidores / Inscritos *", min_value=0, step=1000, key="cad_seg")
            
            rede_detectada = "Instagram"
            if "youtube.com" in link_cria.lower() or "youtu.be" in link_cria.lower(): rede_detectada = "YouTube"
            elif "tiktok.com" in link_cria.lower(): rede_detectada = "TikTok"
            
            st.markdown(f"#### 📊 Calculadora Automática de Engajamento para: **{rede_detectada.upper()}**")
            e1, e2, e3 = st.columns(3)
            views = e1.number_input("Número de Views (Total acumulado)", min_value=0, step=100, key="val_v")
            curtidas = e2.number_input("Número de Curtidas (Total acumulado)", min_value=0, step=100, key="val_c")
            comentarios = e3.number_input("Número de Comentários (Total acumulado)", min_value=0, step=10, key="val_m")
            
            st.markdown("#### 🏡 Informações Complementares (Opcional)")
            co1, co2, co3 = st.columns(3)
            cidade_cria = co1.text_input("Cidade")
            uf_cria = co2.text_input("UF (Ex: SP, RJ)")
            obs_cria = co3.text_input("Status / OBS.", value="ATIVA")
            
            if st.button("💾 Salvar Cadastro no Excel", key="btn_salvar_novo"):
                if not nome_cria or not link_cria or seg_cria == 0:
                    st.error("Por favor, preencha Nome, Link e Seguidores!")
                else:
                    v_atual, c_atual, m_atual = int(views), int(curtidas), int(comentarios)
                    total_interacoes = v_atual + c_atual + m_atual
                    taxa_calculada = (total_interacoes / seg_cria) * 100 if total_interacoes > 0 else 0.0
                    
                    if seg_cria <= 10000: p_man = "Nano / UGC"
                    elif seg_cria <= 199000: p_man = "Micro"
                    elif seg_cria <= 999000: p_man = "Mid-tier"
                    else: p_man = "Macro"
                    
                    payload = {
                        'nome': nome_cria, 'link_principal': link_cria, 'seguidores': seg_cria,
                        'cidade': cidade_cria, 'uf': uf_cria.upper(), 'obs': obs_cria,
                        'rede_detectada': rede_detectada, 'engajamento': taxa_calculada,
                        'porte_manual': p_man, 'rede_manual': rede_detectada,
                        'teve_engajamento': total_interacoes > 0, 'v': v_atual, 'c': c_atual, 'm': m_atual
                    }
                    
                    if salvar_novo_registro_excel(payload):
                        st.toast("Salvo no Excel com sucesso!", icon="❤️")
                        st.balloons()
                        time.sleep(1.2)
                        st.rerun()

        with tab3:
            st.subheader("📈 Relatórios de Auditoria")
            col_r1, col_r2, col_r3 = st.columns(3)
            col_r1.metric("Total de Criadoras", len(df_base))
            if "PERFIL" in df_base.columns:
                col_r2.metric("Total Micro-influenciadoras", len(df_base[df_base["PERFIL"] == "Micro"]))
                col_r3.metric("Total Nano / UGC", len(df_base[df_base["PERFIL"] == "Nano / UGC"]))

elif st.session_state.page == 'visitante':
    df_base = carregar_e_processar_dados()
    st.title("🔍 Espaço de Planejamento de Mídia (Visitante)")
    if st.button("⬅ Voltar para Início"): st.session_state.page = 'entrada'; st.rerun()
    
    v1, v2, v3 = st.columns(3)
    p_vis = v1.multiselect("Filtrar por Perfil:", ["Nano / UGC", "Micro", "Mid-tier", "Macro", "Hero"])
    r_vis = v2.multiselect("Filtrar por Rede Principal:", ["Instagram", "YouTube", "TikTok", "Kwai"])
    uf_ops = sorted(list(df_base["UF"].dropna().astype(str).unique())) if "UF" in df_base.columns else []
    u_vis = v3.multiselect("Filtrar por Região (UF):", uf_ops)
    
    df_v = df_base.copy()
    if p_vis: df_v = df_v[df_v["PERFIL"].isin(p_vis)]
    if r_vis: df_v = df_v[df_v["PRINCIPAL REDE"].isin(r_vis)]
    if u_vis: df_v = df_v[df_v["UF"].isin(u_vis)]
    
    if not df_v.empty:
        df_mask = df_v.reset_index(drop=True)
        df_mask["CRIADORA"] = [f"Criadora {i+1}" for i in range(len(df_mask))]
        
        colunas_bloqueadas = ["E-mail de contato", "Telefone", "OBS"]
        for col in df_mask.columns:
            if col in colunas_bloqueadas:
                df_mask.loc[5:, col] = "🔒 Reservado Equipe"
        
        df_mask.loc[5:, "Link Instagram"] = ""
        df_mask.loc[5:, "Link YouTube"] = ""
        df_mask.loc[5:, "Link TikTok"] = ""
        
        st.warning("🔒 Links de direcionamento e dados de contato estão abertos apenas para os 5 primeiros perfis demonstrativos.")
        
        colunas_exibicao = ["CRIADORA", "PERFIL", "PRINCIPAL REDE", "CIDADE", "UF", "Link Instagram", "Link YouTube", "Link TikTok"]
        
        df_mask_render = df_mask[colunas_exibicao].copy()
        for c_str in colunas_exibicao:
            df_mask_render[c_str] = df_mask_render[c_str].fillna("").astype(str)

        st.dataframe(
            df_mask_render, use_container_width=True, hide_index=True,
            column_config={
                "Link Instagram": st.column_config.LinkColumn("📸 INSTAGRAM", display_text="Simular Link"),
                "Link YouTube": st.column_config.LinkColumn("📺 YOUTUBE", display_text="Simular Link"),
                "Link TikTok": st.column_config.LinkColumn("🎵 TIKTOK", display_text="Simular Link")
            }
        )
        
        csv = df_mask_render.to_csv(index=False).encode('utf-8')
        st.download_button("📥 Exportar Plano de Mídia Filtrado (CSV)", csv, "plano_de_midia_home.csv", "text/csv")